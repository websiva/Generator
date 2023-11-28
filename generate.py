from pathlib import Path
from openpyxl import load_workbook
import os, subprocess, fitz, time

generic_count=None
creo_version=input("Creo Version '2.0' or '4.0'")
creo_sub_version=input("Creo sub version ('M100' or 'M150')")
creo_run=input("Creo running mode 'UI' or 'bg'")
exe_file_location= os.getcwd()
print(f"Current directory: {exe_file_location}")
CAD_directory=exe_file_location.replace("\\", "/")

def read_file_content(file_path):
    with open(file_path, 'r') as file:
        return file.read()

def replace_placeholders(content, replacements):
    for placeholder, value in replacements.items():
        content = content.replace(placeholder, value)
    return content

def directory():
    locations = [
    os.path.join(exe_file_location, 'output', '2dpdf'),
    os.path.join(exe_file_location, 'output', '3dpdf'),
    os.path.join(exe_file_location, 'output', 'json'),
    os.path.join(exe_file_location, 'output', 'tiff'),
    os.path.join(exe_file_location, 'output', 'jpeg'),
    os.path.join(exe_file_location, 'output', 'stp'),
    os.path.join(exe_file_location, 'output', 'asm'),
    os.path.join(exe_file_location, 'output', 'x_t'),
    os.path.join(exe_file_location, 'output', 'igs'),
    os.path.join(exe_file_location, 'output', 'sat'),
    os.path.join(exe_file_location, 'output', 'simplified'),
    os.path.join(exe_file_location, 'output', 'sld'),
    os.path.join(exe_file_location, 'output', 'obj'),
    os.path.join(exe_file_location, 'output', 'asm','jpeg'),
    os.path.join(exe_file_location, 'output', 'x_t','jpeg'),
    os.path.join(exe_file_location, 'output', 'sat','jpeg'),
    os.path.join(exe_file_location, 'output', 'igs','jpeg'),
    os.path.join(exe_file_location, 'output', 'obj','jpeg')
             ]

    for location in locations:
        Path(location).mkdir(parents=True, exist_ok=True)
        
def list_and_choose_files(directory):
    txt_files = [file for file in os.listdir(directory) if file.endswith(".txt")]

    if not txt_files:
        print("No .txt files found in the directory.")
        return None

    print("List of .txt files:")
    for i, txt_file in enumerate(txt_files, start=1):
        print(f"{i}. {txt_file}")

    try:
        selected_number = int(input("Enter the number of the mid template: "))
        if 1 <= selected_number <= len(txt_files):
            selected_file = txt_files[selected_number - 1]
            print(f"You selected: {selected_file}")
            return selected_file
        else:
            print("Invalid input. Please enter a valid file number.")
    except ValueError:
        print("Invalid input. Please enter a number.")

    return None

def create_trail_into_txt():
    directory()
    txt_output_dir=f"{exe_file_location}\\final_trail.txt"
    
    file_name_input = f"{exe_file_location}\\1.xlsx"
    load_wb_input = load_workbook(file_name_input, data_only=True)
    load_ws_input = load_wb_input['Sheet1']
    row_count_input = load_ws_input.max_row  
    global generic_count 
    generic_count = row_count_input - 9
    column_count_input = load_ws_input.max_column
    mid_template=list_and_choose_files("E:\\001_new_generator")
    CAD_name=load_ws_input.cell(row=1,column=2).value
    DRW_name=load_ws_input.cell(row=2,column=2).value
    
    beginning_lines_content = read_file_content("E:\\001_new_generator\\0001_top_template.txt")
    mid_template_content = read_file_content(f"E:\\001_new_generator\\{mid_template}")
    bottom_line_content = read_file_content("E:\\001_new_generator\\01_bottom_template.txt")

    with open(txt_output_dir, 'w') as txt_output_file:
        starting_row_number = 10
        for j in range(1, generic_count + 1):
            part_number=str(load_ws_input.cell(row=starting_row_number,column=1).value)
            replacements = {
                '<% SOURCE MODEL %>': CAD_name,
                '<% DRAWING %>': DRW_name,
                '<% V:PART_NUMBER %>': part_number,
                '<% CAD_directory %>': CAD_directory
                }
            if(j==1):
                beginning_lines_content = replace_placeholders(beginning_lines_content, replacements)
                txt_output_file.write(beginning_lines_content + '\n') 
            lines_to_insert = []
            parameter_row_number = 7
            column_number = 2 
            for column_number in range(2, column_count_input + 1):
                parameter_value = str(load_ws_input.cell(row=starting_row_number, column=column_number).value)
                parameter_name = str(load_ws_input.cell(row=parameter_row_number, column=column_number).value)
                parameter_len = str(parameter_value)
                if parameter_len != "None":
                    lines_to_insert.extend(["#Modify", f"#{parameter_name}", parameter_value])  
            paragraph_value_4 = replace_placeholders(mid_template_content, replacements)
            if "<% PARAMETERS %>" in paragraph_value_4:
                paragraph_value_4= paragraph_value_4.replace('<% PARAMETERS %>', "\n".join(lines_to_insert))
            txt_output_file.write(paragraph_value_4+ '\n')
            starting_row_number=starting_row_number+1
        txt_output_file.write(bottom_line_content + '\n')

def exe_file():
    output_directory = os.path.join(exe_file_location, "output")
    subprocess.run(["rmdir", "/s", "/q", output_directory], shell=True)
    subprocess.run(["DEL", os.path.join(exe_file_location, "*trail.txt*")], shell=True)
    
    trail_creation_args = create_trail_into_txt()
    if trail_creation_args:
        subprocess.run(trail_creation_args)

    if(creo_version=="4.0"):
        creo_path = fr'C:\Program Files\PTC\Creo {creo_version}\{creo_sub_version}\Parametric\bin\parametric.exe'
    else:
        creo_path = fr'C:\Program Files\PTC\Creo {creo_version}\Parametric\bin\parametric.exe'
    creo_command = [
        creo_path,
        "-g:no_graphics",
        "-batch_mode",
        "-i:rpc_input",
        "final_trail.txt"
    ]

    try:
        if creo_run == "UI":
            subprocess.run([creo_path, "final_trail.txt"], shell=True)
        else:
            subprocess.run(creo_command, check=True, shell=True)
    except subprocess.CalledProcessError as e:
        print(f"Error running Creo command: {e}")

def main():
    exe_file()

if __name__ == "__main__":
    main()
